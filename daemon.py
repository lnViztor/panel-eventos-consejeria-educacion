import win32api
import win32con
import win32gui
import win32process
import time
import warnings
import psutil
from ctypes import *
from openpyxl import load_workbook
from pathlib import Path

# pip install pywin32
# pip install openpyxl
# pip install psutil

warnings.simplefilter("ignore")

#
# Device change events (WM_DEVICECHANGE wParam)
#
DBT_DEVICEARRIVAL = 0x8000
DBT_DEVICEQUERYREMOVE = 0x8001
DBT_DEVICEQUERYREMOVEFAILED = 0x8002
DBT_DEVICEMOVEPENDING = 0x8003
DBT_DEVICEREMOVECOMPLETE = 0x8004
DBT_DEVICETYPESSPECIFIC = 0x8005
DBT_CONFIGCHANGED = 0x0018

#
# type of device in DEV_BROADCAST_HDR
#
DBT_DEVTYP_OEM = 0x00000000
DBT_DEVTYP_DEVNODE = 0x00000001
DBT_DEVTYP_VOLUME = 0x00000002
DBT_DEVTYPE_PORT = 0x00000003
DBT_DEVTYPE_NET = 0x00000004

#
# media types in DBT_DEVTYP_VOLUME
#
DBTF_MEDIA = 0x0001
DBTF_NET = 0x0002

WORD = c_ushort
DWORD = c_ulong

# Para ocultar la barra de tareas
SW_HIDE = 0
SW_SHOW = 9  # o 5
user32 = windll.user32
# Buscar la barra de tareas principal
h_taskbar = user32.FindWindowW("Shell_TrayWnd", None)


class DEV_BROADCAST_HDR (Structure):
    _fields_ = [
        ("dbch_size", DWORD),
        ("dbch_devicetype", DWORD),
        ("dbch_reserved", DWORD)
    ]


class DEV_BROADCAST_VOLUME (Structure):
    _fields_ = [
        ("dbcv_size", DWORD),
        ("dbcv_devicetype", DWORD),
        ("dbcv_reserved", DWORD),
        ("dbcv_unitmask", DWORD),
        ("dbcv_flags", WORD)
    ]


def drive_from_mask(mask):
    n_drive = 0
    while 1:
        if (mask & (2 ** n_drive)):
            return n_drive
        else:
            n_drive += 1


def get_chrome_hwnds(window_name):
    def callback(hwnd, hwnds):
        if win32gui.IsWindowVisible(hwnd) and win32gui.IsWindowEnabled(hwnd):
            name = win32gui.GetWindowText(hwnd)
            # print(name)
            if "Tablero" in name:
                hwnds.append(hwnd)
        return True

    hwnds = []
    win32gui.EnumWindows(callback, hwnds)
    return hwnds


def hide_taskbar():
    # buscar el handle en el momento de la llamada
    h_taskbar = user32.FindWindowW("Shell_TrayWnd", None)
    if h_taskbar:
        user32.ShowWindow(h_taskbar, SW_HIDE)
    # también intentar esconder la bandeja secundaria (multi-monitor)
    h_secondary = user32.FindWindowW("Shell_SecondaryTrayWnd", None)
    if h_secondary:
        user32.ShowWindow(h_secondary, SW_HIDE)
    return True


def show_taskbar():
    if h_taskbar:
        user32.ShowWindow(h_taskbar, SW_SHOW)


class Notification:

    def __init__(self):
        message_map = {
            win32con.WM_DEVICECHANGE: self.onDeviceChange
        }

        wc = win32gui.WNDCLASS()
        hinst = wc.hInstance = win32api.GetModuleHandle(None)
        wc.lpszClassName = "DeviceChangeDemo"
        wc.style = win32con.CS_VREDRAW | win32con.CS_HREDRAW
        wc.hCursor = win32gui.LoadCursor(0, win32con.IDC_ARROW)
        wc.hbrBackground = win32con.COLOR_WINDOW
        wc.lpfnWndProc = message_map
        classAtom = win32gui.RegisterClass(wc)
        style = win32con.WS_OVERLAPPED | win32con.WS_SYSMENU
        self.hwnd = win32gui.CreateWindow(
            classAtom,
            "Device Change Demo",
            style,
            0, 0,
            win32con.CW_USEDEFAULT, win32con.CW_USEDEFAULT,
            0, 0,
            hinst, None
        )

    def onDeviceChange(self, hwnd, msg, wparam, lparam):
        dev_broadcast_hdr = DEV_BROADCAST_HDR.from_address(lparam)

        if wparam == DBT_DEVICEARRIVAL:
            print("USB conectado")

            if dev_broadcast_hdr.dbch_devicetype == DBT_DEVTYP_VOLUME:

                dev_broadcast_volume = DEV_BROADCAST_VOLUME.from_address(
                    lparam)
                drive_letter = drive_from_mask(
                    dev_broadcast_volume.dbcv_unitmask)
                unidad = chr(ord("A") + drive_letter)
                print("Detectada unidad: ", unidad)
                # comprobar xlsx
                try:
                    my_file = Path(unidad + "://eventos.xlsx")
                    if my_file.is_file():
                        my_file
                        wb = load_workbook(
                            unidad + "://eventos.xlsx", read_only=True)
                        ws = wb['Hoja1']
                        # leer xlsx
                        eventos = []
                        for i in range(2, 12):
                            sala = ws['A'+str(i)].value
                            evento = ws['B'+str(i)].value
                            if evento is not None:
                                eventos.append((sala, evento))
                        print(eventos)

                        divs = ""
                        for evento in eventos:
                            divs = "⏰" + divs + '<div id="linea"><p class="lugarhora">' + "🗒️"+ \
                                evento[0] + '</p><p class="evento">' +\
                                evento[1] + '</p></div>'

                        # guardar html
                        f = open('index.template.html', 'r', encoding='utf-8')
                        contents = f.readlines()
                        f.close()

                        index = 0
                        while index < (len(contents)-1) and not ("-INSERT HERE-" in contents[index]):
                            index = index + 1

                        contents.insert(index, divs)

                        f = open('index.html', 'w', encoding='utf-8')
                        contents = "".join(contents)
                        f.write(contents)
                        f.close()
                        # cerrar chrome si está abierto
                        # cerrar ventanas
                        for hwnd in get_chrome_hwnds("Tablero de Entrada"):
                            win32gui.SendMessage(hwnd, win32con.WM_CLOSE, 0, 0)
                            time.sleep(1)
                        PROCNAME = 'chrome.exe'
                        # cerrar todos los procesos
                        # for proc in psutil.process_iter():
                        #    if proc.name() == PROCNAME:
                        #        p = psutil.Process(proc.pid)
                        #        if not 'SYSTEM' in p.username():
                        #            proc.kill()
                        # abrir chrome pantalla completa
                        time.sleep(1)
                        p = psutil.Popen(["C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
                                          "--kiosk",
                                          "-fullscreen",
                                          "file:\\C:\\Users\\vvso01.JCLM\\Desktop\\panel\\index.html"])
                        wb.close()
                        print("Ocultando barra de tareas...")
                        hide_taskbar()
                        input("Barra oculta. Pulsa Enter para volver a mostrarla...")
                        show_taskbar()
                    else:
                        print('No contiene el archivo eventos.xlsx')
                except Exception as e:
                    print("Error leyendo unidad/archivo")
                    print(e)
        return 1


if __name__ == '__main__':
    w = Notification()
    win32gui.PumpMessages()
